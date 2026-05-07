"""enrichment_rbook.py — vAuto rBook competitive-set scraper (production).

Discovered DOM pattern via spike on 2026-05-06:
  - Click count link in section anchored by 'Provision rBook - Live Retail Market View'
  - Competitive set rows render IN the appraisal app's shadow root (NOT a popup)
  - Each row is a div with class containing 'row', containing year/make/model,
    VIN (17 chars), color/interior, $price, mileage, days, dealer name
  - Table is virtualized — only ~12 rows visible at a time (scroll-to-load
    not implemented in this version; returns visible window)

Returns dict:
  {
    'rows': [{vin, price, mileage, days_on_lot, color, interior, dealer}, ...],
    'count_text': '(200)',  # total in competitive set per vAuto
    'panel_found': true,
    'stocking_report': {demand, interest, volume, days_supply, availability}
  }
"""

from __future__ import annotations
import time
import re


# Module-level competition-set JSON capture. The listener accumulates across
# scrapes (Playwright's ctx.on doesn't dedupe handler closures perfectly), so
# binding to a module-level dict means whichever listener fires updates the
# same target — no more "captured" log + "never fired" check stuck on stale
# closures. Each scrape resets _COMPETITION_CAPTURE['value'] = None at start.
_COMPETITION_CAPTURE = {'value': None}


def _on_competition_response(resp):
    """Module-level Playwright response listener — writes any
    /api/competition/vehicles POST body into _COMPETITION_CAPTURE."""
    try:
        url = resp.url or ''
        if '/api/competition/vehicles' in url and resp.request.method == 'POST':
            if resp.status == 200:
                body = resp.json()
                _COMPETITION_CAPTURE['value'] = body
                try:
                    size = len(resp.body() or b'')
                except Exception:
                    size = 0
                print(f'  [rbook] captured /api/competition/vehicles JSON '
                      f'({size:,} bytes)', flush=True)
    except Exception as e:
        print(f'  [rbook] response-capture error: {e}', flush=True)


# JS — find the rBook count link and return its rect (NO click here).
# The Python worker uses page.mouse.click(cx, cy) for a TRUSTED user event
# that actually triggers vAuto's popup handlers. Synthetic JS clicks via
# el.click() are reported as successful but don't fire the real handler.
JS_FIND_RBOOK_TARGET = r"""
() => {
    function allEls(root) {
        const out = [];
        function walk(node) {
            if (!node || !node.querySelectorAll) return;
            try {
                node.querySelectorAll('*').forEach(el => {
                    out.push(el);
                    if (el.shadowRoot) walk(el.shadowRoot);
                });
            } catch (e) {}
        }
        walk(root);
        return out;
    }
    const all = allEls(document);
    let anchor = null;
    for (const el of all) {
        const t = (el.textContent || '').trim();
        if (t.length > 200) continue;
        if (/Provision\s+rBook|Live\s+Retail\s+Market|Live\s+Market\s+View/i.test(t)) {
            anchor = el; break;
        }
    }
    if (!anchor) return {result: 'no_anchor'};
    let container = anchor;
    for (let i = 0; i < 8 && container; i++) {
        const t = (container.textContent || '').toLowerCase();
        if (t.indexOf('vehicles') >= 0) {
            const subs = allEls(container);
            for (const el of subs) {
                const txt = (el.textContent || '').trim();
                if (/^\(?\d+\)?$/.test(txt) && el.offsetParent) {
                    let p = el;
                    for (let j = 0; j < 5 && p; j++) {
                        const ptxt = (p.textContent || '').toLowerCase();
                        if (ptxt.indexOf('vehicles') >= 0) {
                            el.scrollIntoView?.({block: 'center', behavior: 'instant'});
                            const r = el.getBoundingClientRect();
                            return {
                                result: 'found',
                                countText: txt,
                                tag: el.tagName,
                                cls: (el.className || '').toString().substring(0, 80),
                                cx: r.left + r.width / 2,
                                cy: r.top + r.height / 2,
                                w: r.width,
                                h: r.height,
                            };
                        }
                        p = p.parentElement;
                    }
                }
            }
            break;
        }
        container = container.parentElement;
    }
    return {result: 'no_count'};
}
"""


# JS — find the Excel-export click target and return its bounding rect.
# Target the INNER icon (excel-div / excel-nav) NOT the outer ids-tooltip-trigger
# span — the span has class "w-full" (spans the toolbar) so clicking its center
# lands on whitespace between icons. The icon itself is the visible click target
# the user actually hits.
JS_FIND_EXCEL_TARGET = r"""
() => {
    function allEls(root, out) {
        out = out || [];
        if (!root || !root.querySelectorAll) return out;
        try {
            root.querySelectorAll('*').forEach(el => {
                out.push(el);
                if (el.shadowRoot) allEls(el.shadowRoot, out);
            });
        } catch (e) {}
        return out;
    }
    const all = allEls(document);

    // Try in order: I.excel-nav (the actual icon), DIV.excel-div (the
    // icon's tight container), then SPAN.ids-tooltip-trigger (last resort,
    // wide). Pick whichever has a SMALL non-zero rect — that's the icon.
    let candidates = [];
    for (const el of all) {
        if (!el.offsetParent) continue;
        const cls = (el.className || '').toString();
        if (/\bexcel-nav\b/.test(cls)) candidates.push({el, kind: 'excel-nav'});
        else if (/\bexcel-div\b/.test(cls)) candidates.push({el, kind: 'excel-div'});
    }

    if (!candidates.length) return {result: 'no_excel_target'};

    // Prefer the candidate with the SMALLEST visible width — that's the icon
    candidates = candidates.map(c => {
        const r = c.el.getBoundingClientRect();
        return {...c, rect: r, area: r.width * r.height};
    }).filter(c => c.rect.width > 0 && c.rect.height > 0)
      .sort((a, b) => a.area - b.area);

    if (!candidates.length) return {result: 'no_visible_target'};

    const chosen = candidates[0];
    chosen.el.scrollIntoView?.({block: 'center', behavior: 'instant'});
    // Re-fetch rect after scrollIntoView
    const r = chosen.el.getBoundingClientRect();

    // Climb up to 4 ancestors and OR-fold class names — vAuto puts the
    // `in-active` state on the parent excel-div, not always on the I.
    let activityFlags = '';
    let node = chosen.el;
    for (let i = 0; i < 4 && node; i++) {
        activityFlags += ' ' + ((node.className || '').toString());
        node = node.parentElement;
    }
    const isInactive = /\bin-active\b/i.test(activityFlags)
                    || /\bdisabled\b/i.test(activityFlags);

    return {
        result: 'found',
        tag: chosen.el.tagName,
        cls: (chosen.el.className || '').toString().substring(0, 100),
        kind: chosen.kind,
        is_inactive: isInactive,
        flags: activityFlags.substring(0, 200),
        x: r.left,
        y: r.top,
        cx: r.left + r.width / 2,
        cy: r.top + r.height / 2,
        w: r.width,
        h: r.height,
        candidates_count: candidates.length,
    };
}
"""

# Legacy JS click — kept for completeness but NOT used. Synthetic clicks
# don't trigger download events because browsers require a real user gesture.
JS_CLICK_EXPORT_EXCEL = r"""
() => {
    function allEls(root, out) {
        out = out || [];
        if (!root || !root.querySelectorAll) return out;
        try {
            root.querySelectorAll('*').forEach(el => {
                out.push(el);
                if (el.shadowRoot) allEls(el.shadowRoot, out);
            });
        } catch (e) {}
        return out;
    }
    const all = allEls(document);

    // Discovered via record_clicks.py spike on bid 928:
    //   The ACTUAL click handler is on <SPAN class="ids-tooltip-trigger ..."> —
    //   the wrapper around <DIV class="excel-div"> with <I class="excel-nav">.
    //   Need to click the ids-tooltip-trigger SPAN, not its inner div, for
    //   the download handler to fire correctly.

    // Step 1: locate the .excel-div
    let excelDiv = null;
    for (const el of all) {
        if (!el.offsetParent) continue;
        const cls = (el.className || '').toString();
        if (/\bexcel-div\b/.test(cls)) {
            excelDiv = el;
            break;
        }
    }
    if (!excelDiv) {
        // Fallback: locate via <I class="excel-nav">
        for (const el of all) {
            if (!el.offsetParent) continue;
            const cls = (el.className || '').toString();
            if (/\bexcel-nav\b/.test(cls)) {
                excelDiv = el.closest('[class*="excel-div"]') || el.parentElement || el;
                break;
            }
        }
    }
    if (!excelDiv) return {result: 'no_excel_div'};

    // Step 2: walk up to the ids-tooltip-trigger SPAN — the actual click handler.
    // The click target the user actually hits in the recording is this span.
    let target = excelDiv;
    let depth = 0;
    while (target && target.parentNode && depth < 6) {
        const cls = (target.className || '').toString();
        if (/\bids-tooltip-trigger\b/.test(cls)) break;
        // parentNode handles ShadowRoot which doesn't have parentElement
        target = target.parentNode.host || target.parentNode;
        depth++;
    }
    if (!target || target.nodeType !== 1) target = excelDiv;

    try { target.scrollIntoView?.({block: 'center'}); } catch (e) {}

    // Dispatch a full click event sequence (some vAuto handlers need
    // mousedown+mouseup, not just .click()).
    try {
        const r = target.getBoundingClientRect();
        const cx = r.left + r.width / 2;
        const cy = r.top + r.height / 2;
        const opts = { bubbles: true, cancelable: true, view: window,
                       clientX: cx, clientY: cy, button: 0 };
        target.dispatchEvent(new MouseEvent('mousedown', opts));
        target.dispatchEvent(new MouseEvent('mouseup', opts));
        target.dispatchEvent(new MouseEvent('click', opts));
    } catch (e) {
        // Last resort
        target.click();
    }
    return {result: 'clicked',
            tag: target.tagName,
            cls: (target.className || '').toString().substring(0, 100),
            via: 'ids-tooltip-trigger'};
}
"""


# JS — toggle "Show My Vehicle" — verifier-exact pattern. Critical: without
# this, vAuto hides the subject + sometimes hides comp rows entirely.
JS_TOGGLE_SHOW_MY_VEHICLE = r"""
() => {
    function allEls(root, out) {
        out = out || [];
        if (!root || !root.querySelectorAll) return out;
        try {
            root.querySelectorAll('*').forEach(el => {
                out.push(el);
                if (el.shadowRoot) allEls(el.shadowRoot, out);
            });
        } catch (e) {}
        return out;
    }
    const all = allEls(document);
    let target = null;
    for (const el of all) {
        const t = (el.textContent || '').trim();
        if (t.length > 60) continue;
        if (/^show\s+my\s+vehicle$/i.test(t)) { target = el; break; }
    }
    if (!target) return 'no_toggle';
    let container = target;
    for (let i = 0; i < 5 && container; i++) {
        const toggle = container.querySelector(
            'input[type="checkbox"], [role="switch"], ids-switch, .k-switch, button[aria-pressed]');
        if (toggle) {
            const checked = toggle.checked
                || toggle.getAttribute('aria-pressed') === 'true'
                || toggle.getAttribute('aria-checked') === 'true';
            if (!checked) { toggle.click(); return 'toggled_on'; }
            return 'already_on';
        }
        container = container.parentElement;
    }
    target.click();
    return 'label_clicked';
}
"""


# JS — extract competitive set rows + stocking report from the appraisal
# shadow root after the rBook click has populated content.
JS_EXTRACT_RBOOK = r"""
() => {
    const app = document.querySelector('profit-time-guided-appraisal');
    if (!app || !app.shadowRoot) return {error: 'no_appraisal_app'};

    function allEls(root) {
        const out = [];
        function walk(node) {
            if (!node || !node.querySelectorAll) return;
            try {
                node.querySelectorAll('*').forEach(el => {
                    out.push(el);
                    if (el.shadowRoot) walk(el.shadowRoot);
                });
            } catch (e) {}
        }
        walk(root);
        return out;
    }
    const all = allEls(app.shadowRoot);

    // Verifier-pattern row finder: discover all VINs in shadow root, then
    // for each VIN find the SMALLEST containing element with both that VIN
    // AND a $ amount. Walk up if needed to include the $ cell.
    // No text-length window — works regardless of how the row's text is
    // distributed across child elements (different cars render differently).
    const VIN_RE = /\b([A-HJ-NPR-Z0-9]{17})\b/g;

    // Step 1: discover every unique VIN in the entire shadow root. The
    // top-level html scan is faster than per-element matching when there
    // are 1000s of elements.
    const fullText = (app.shadowRoot.innerHTML || '');
    // Strip HTML tags for VIN regex (avoids matching id="..." attribute values)
    const visibleText = fullText.replace(/<[^>]+>/g, ' ').replace(/&[a-z]+;/g, ' ');
    const vinSet = new Set();
    let vinM;
    while ((vinM = VIN_RE.exec(visibleText)) !== null) {
        vinSet.add(vinM[1]);
    }
    const allVins = [...vinSet];

    // Step 2: for each VIN, find the smallest containing element that ALSO
    // contains a dollar amount. That's the row.
    const rows = [];
    const seenContainers = new Set();
    for (const vin of allVins) {
        let bestEl = null;
        let bestLen = 999999;
        for (const el of all) {
            const txt = el.textContent || '';
            if (txt.length > 5000) continue;       // skip whole-panel wrappers
            if (txt.indexOf(vin) < 0) continue;
            if (!txt.includes('$')) continue;
            if (txt.length < bestLen) {
                bestEl = el;
                bestLen = txt.length;
            }
        }
        if (!bestEl) continue;
        if (seenContainers.has(bestEl)) continue;
        seenContainers.add(bestEl);

        const txt = (bestEl.textContent || '').trim();

        // Parse fields from text content
        const priceM   = txt.match(/\$\s*([\d,]+)/);
        const milesM   = txt.match(/([\d,]+)\s*mi\b/i);
        const colorM   = txt.match(/Color\s*:?\s*([A-Za-z][A-Za-z\s\-]{1,20})/i);
        const interM   = txt.match(/Interior\s*:?\s*([A-Za-z][A-Za-z\s\-]{1,20})/i);
        const ymmM     = txt.match(/(\d{4})\s+([A-Z][A-Za-z\-]+)\s+([A-Z][A-Za-z\s\-]+?)(?=VIN|\s{2,}|Color|$)/);

        let dealer = null;
        if (priceM) {
            const after = txt.substring(priceM.index + priceM[0].length);
            const dealerM = after.match(/([A-Z][A-Z\s&\.\-,]{6,80}[A-Z])/);
            if (dealerM) dealer = dealerM[1].trim().replace(/\s+/g, ' ');
        }

        let days = null;
        if (milesM) {
            const tail = txt.substring(milesM.index + milesM[0].length,
                                       milesM.index + milesM[0].length + 80);
            const daysM = tail.match(/\.?\s*(\d{1,3})\b/);
            if (daysM) days = parseInt(daysM[1], 10);
        }

        rows.push({
            vin, raw_text: txt.substring(0, 400),
            text_len: txt.length,
            year: ymmM ? parseInt(ymmM[1], 10) : null,
            make: ymmM ? ymmM[2] : null,
            model: ymmM ? ymmM[3].trim() : null,
            price: priceM ? parseInt(priceM[1].replace(/,/g, ''), 10) : null,
            mileage: milesM ? parseInt(milesM[1].replace(/,/g, ''), 10) : null,
            days_on_lot: days,
            color: colorM ? colorM[1].trim() : null,
            interior: interM ? interM[1].trim() : null,
            dealer: dealer
        });
    }

    // Stocking Report Card (single row at top of rBook with letter grades)
    let stocking = null;
    for (const el of all) {
        const t = (el.textContent || '').trim();
        if (t.indexOf('Stocking Report') < 0) continue;
        if (t.length > 500) continue;
        const get = (label) => {
            const re = new RegExp(label + '\\s*([A-D][+\\-]?|F)', 'i');
            const m = t.match(re);
            return m ? m[1] : null;
        };
        stocking = {
            demand:       get('Demand'),
            interest:     get('Interest'),
            volume:       get('Volume'),
            days_supply:  get('Days Supply'),
            availability: get('Availability'),
            raw_text:     t.substring(0, 300),
        };
        break;
    }

    // Count text for the rBook section (e.g. "(200)")
    let countText = null;
    for (const el of all) {
        const t = (el.textContent || '').trim();
        if (/^\(?\d+\)?$/.test(t) && el.offsetParent) {
            const parentText = (el.parentElement?.textContent || '').toLowerCase();
            if (parentText.indexOf('vehicles') >= 0 || parentText.indexOf('rbook') >= 0) {
                countText = t; break;
            }
        }
    }

    return {rows, stocking_report: stocking, count_text: countText,
            n_visible: rows.length};
}
"""


def _parse_xlsx(xlsx_path) -> list[dict]:
    """Parse vAuto's Comparable Vehicles export. Columns vary slightly by
    region/version; we use header-name detection + flexible field matching.
    Returns a list of row dicts compatible with enrichment_rbook.scrape's
    output rows."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        print('  [rbook] openpyxl not available — Excel parse skipped', flush=True)
        return []

    try:
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
        ws = wb.active
        # First row may be a title; scan for the actual header row by looking
        # for a row that contains 'Year' or 'VIN' or 'Price'.
        header_row_idx = None
        rows_iter = list(ws.iter_rows(values_only=True))
        for i, row in enumerate(rows_iter):
            joined = ' '.join(str(c) for c in row if c is not None).lower()
            if ('year' in joined or 'vin' in joined) and 'price' in joined:
                header_row_idx = i
                break
        if header_row_idx is None:
            print(f'  [rbook] xlsx: no header row found (rows={len(rows_iter)})', flush=True)
            return []

        headers = [str(c).strip() if c else '' for c in rows_iter[header_row_idx]]
        print(f'  [rbook] xlsx headers: {headers}', flush=True)

        # Map column names to our schema
        def find_col(*candidates):
            for cand in candidates:
                for i, h in enumerate(headers):
                    if cand.lower() in h.lower():
                        return i
            return None

        col_year   = find_col('year')
        col_make   = find_col('make')
        col_model  = find_col('model')
        col_trim   = find_col('trim', 'series')
        col_vin    = find_col('vin')
        col_color  = find_col('exterior color', 'color')
        col_price  = find_col('list price', 'price', 'asking')
        col_miles  = find_col('odometer', 'mileage', 'miles')
        col_days   = find_col('days on lot', 'days', 'age')
        col_dealer = find_col('dealer', 'seller', 'merchant')
        col_dist   = find_col('distance')

        rows_out = []
        seen_vins = set()
        for raw_row in rows_iter[header_row_idx + 1:]:
            if not raw_row or all(c is None for c in raw_row):
                continue
            def cell(idx):
                if idx is None or idx >= len(raw_row): return None
                v = raw_row[idx]
                return v if v not in (None, '') else None

            def to_int(x):
                if x is None: return None
                try:
                    if isinstance(x, (int, float)): return int(x)
                    s = ''.join(ch for ch in str(x) if ch.isdigit())
                    return int(s) if s else None
                except Exception: return None

            vin = cell(col_vin)
            if vin: vin = str(vin).strip().upper()
            if vin in seen_vins: continue
            if vin: seen_vins.add(vin)

            rows_out.append({
                'vin':         vin,
                'year':        to_int(cell(col_year)),
                'make':        str(cell(col_make) or '').strip() or None,
                'model':       str(cell(col_model) or '').strip() or None,
                'trim':        str(cell(col_trim) or '').strip() or None,
                'color':       str(cell(col_color) or '').strip() or None,
                'price':       to_int(cell(col_price)),
                'mileage':     to_int(cell(col_miles)),
                'days_on_lot': to_int(cell(col_days)),
                'dealer':      str(cell(col_dealer) or '').strip() or None,
                'distance':    to_int(cell(col_dist)),
            })
        return rows_out
    except Exception as e:
        print(f'  [rbook] xlsx parse error: {e}', flush=True)
        return []


def _parse_competition_json(body: dict) -> list[dict]:
    """Convert vAuto's /api/competition/vehicles response into our row schema.
    Each competitor has VIN, price, mileage, days, dealer, distance, etc."""
    rows = []
    if not body:
        return rows
    competitors = body.get('competitiveSetVehicles') or []
    seen_vins = set()
    for v in competitors:
        if not isinstance(v, dict): continue
        vin = (v.get('vin') or '').strip().upper() or None
        if vin and vin in seen_vins: continue
        if vin: seen_vins.add(vin)

        # vehicleTitle is "2025 Audi Q8 55 Prestige quattro" — split into pieces
        title = (v.get('vehicleTitle') or '').strip()
        title_parts = title.split(None, 3)
        year = None
        try: year = int(title_parts[0]) if title_parts else None
        except (ValueError, IndexError): year = None
        make  = title_parts[1] if len(title_parts) > 1 else None
        model = title_parts[2] if len(title_parts) > 2 else None
        trim  = title_parts[3] if len(title_parts) > 3 else None

        rows.append({
            'vin':                  vin,
            'year':                 year,
            'make':                 make,
            'model':                model,
            'trim':                 trim,
            'color':                v.get('exteriorBaseColor'),
            'interior':             v.get('interiorDescription'),
            'price':                v.get('price'),
            'effective_price':      v.get('effectivePrice'),
            'mileage':              v.get('odometer'),
            'days_on_lot':          v.get('daysInInventory'),
            'dealer':               v.get('sellerName'),
            'dealer_city':          v.get('sellerCity'),
            'dealer_state':         v.get('sellerRegion'),
            'dealer_postal':        v.get('sellerPostalCode'),
            'distance':             v.get('distance'),
            'rank':                 v.get('rank'),
            'pending_sale':         v.get('pendingSale'),
            'is_certified':         v.get('isCertified'),
            'carfax_one_owner':     v.get('carfaxOneOwner'),
            'carfax_clean_title':   v.get('carfaxCleanTitle'),
            'body':                 v.get('body'),
            'engine':               v.get('engine'),
            'transmission':         v.get('transmission'),
            'drivetrain':           v.get('driveTrain'),
            'detail_uri':           v.get('detailUri'),
        })
    return rows


def scrape(page, job: dict) -> dict | None:
    """Click the rBook count link → toggle Show My Vehicle → click Export to
    Excel → parse the downloaded xlsx for ALL rows (not just the virtualized
    visible window). Falls back to in-DOM scrape if Excel export fails.

    Returns dict (see module docstring) or None if rBook never populated.
    """
    # Debug screenshots so we can SEE what state the page is in at each step.
    # Saved to C:\worker\spike_rbook\debug_<bidid>_<step>.png on the VM.
    import os as _os
    DEBUG_DIR = r'C:\worker\spike_rbook'
    try: _os.makedirs(DEBUG_DIR, exist_ok=True)
    except Exception: pass
    bid_id = job.get('bid_id', 'x')
    def _shot(label):
        try:
            page.screenshot(path=_os.path.join(DEBUG_DIR, f'debug_{bid_id}_{label}.png'),
                            full_page=False)
            print(f'  [rbook] screenshot: debug_{bid_id}_{label}.png', flush=True)
        except Exception as e:
            print(f'  [rbook] screenshot failed: {e}', flush=True)

    _shot('00_pre_click')

    # Step 1: poll for rBook count link to appear. Fresh appraisals can take
    # 60-90s for vAuto to fetch rBook data on-demand, so we wait up to 120s.
    # Trusted user-gesture click (page.mouse.click) is required — JS .click()
    # gets reported as success but vAuto's real handlers don't fire.
    rb_target = None
    for attempt in range(40):  # 40 × 3s = 120s
        rb_target = page.evaluate(JS_FIND_RBOOK_TARGET)
        if rb_target and rb_target.get('result') == 'found':
            break
        # Scroll to encourage lazy-loading
        try:
            page.evaluate(r"""() => {
                window.scrollBy(0, 600);
                const app = document.querySelector('profit-time-guided-appraisal');
                if (app && app.shadowRoot) {
                    app.shadowRoot.querySelectorAll('*').forEach(el => {
                        const t = (el.textContent || '');
                        if (/Provision\s+rBook|Live\s+Retail/i.test(t) && el.scrollIntoView) {
                            try { el.scrollIntoView({block: 'center'}); } catch(e) {}
                        }
                    });
                }
            }""")
        except Exception: pass
        time.sleep(3)
        # Log only every 5th attempt to reduce noise
        if attempt < 39 and (attempt % 5 == 4):
            print(f'  [rbook] count-link not ready ({rb_target}) — retry '
                  f'{attempt+1}/40 (waiting for vAuto rBook fetch)',
                  flush=True)

    if not rb_target or rb_target.get('result') != 'found':
        print(f'  [rbook] count-link not found after retries: {rb_target}', flush=True)
        _shot('01_click_failed')
        return None

    expected_count = rb_target.get('countText', '?')
    rb_cx = float(rb_target.get('cx') or 0)
    rb_cy = float(rb_target.get('cy') or 0)
    if rb_cx <= 0 or rb_cy <= 0:
        print(f'  [rbook] count-link offscreen: cx={rb_cx} cy={rb_cy}', flush=True)
        _shot('01_offscreen')
        return None

    # JSON-API path: when the user clicks the rBook count link, vAuto's
    # frontend POSTs to /api/competition/vehicles which returns ALL rows as
    # clean JSON (~180KB, 165+ vehicles). We intercept that response — no
    # Excel button, no xlsx parsing, no popup pagination.
    #
    # IMPORTANT: We use a MODULE-LEVEL capture dict + module-level listener
    # because Playwright's ctx.on('page', ...) accumulates listeners across
    # scrapes — each scrape would create a new local dict, but old listeners
    # (from prior scrapes) still hold references to their old dicts and end
    # up "stealing" the JSON before the new one's wait-loop sees it. By
    # writing to the same module dict, whichever listener fires updates it,
    # and we just reset the dict at the start of each scrape.
    _COMPETITION_CAPTURE['value'] = None
    competition_json = _COMPETITION_CAPTURE  # alias for the rest of this fn

    ctx = page.context
    # Attach listener to existing + future pages. Idempotent — Playwright
    # de-dupes identical handlers. (Even if it didn't, all listeners write
    # to the same dict so duplication doesn't break anything.)
    for p in ctx.pages:
        try: p.on('response', _on_competition_response)
        except Exception: pass
    ctx.on('page', lambda p: p.on('response', _on_competition_response))

    # TRUSTED click via real mouse coordinates — fires the popup AND the
    # /api/competition/vehicles XHR.
    page.mouse.click(rb_cx, rb_cy)
    print(f'  [rbook] mouse.click({int(rb_cx)}, {int(rb_cy)}) on count-link '
          f'{expected_count} ({rb_target.get("tag")} .{(rb_target.get("cls") or "")[:40]})',
          flush=True)

    # Wait up to 90s for the JSON response to land. vAuto USUALLY answers
    # in <5s, but on slow sessions we've seen it take 30-60s — and falling
    # through to the legacy Excel path almost always hangs forever (the
    # excel-nav button stays in-active). 90s here is much cheaper than
    # the 5+ minute Excel hang seen on bid 984.
    json_t_end = time.time() + 90
    while time.time() < json_t_end:
        if competition_json['value'] is not None:
            break
        time.sleep(0.5)

    _shot('01_after_count_click')

    if competition_json['value'] is not None:
        body = competition_json['value']
        rows = _parse_competition_json(body)
        print(f'  [rbook] parsed {len(rows)} rows from competition/vehicles JSON',
              flush=True)
        return {
            'rows': rows,
            'count_text': str(len(rows)),
            'stocking_report': None,
            'panel_found': True,
            'n_visible': len(rows),
            'source': 'competition_api',
        }
    print(f'  [rbook] /api/competition/vehicles never fired — '
          f'falling through to legacy xlsx path', flush=True)
    time.sleep(7)  # let popup settle for legacy path

    # Step 2 — try the Excel export path. Captures ALL rows in the popup.
    rows_from_xlsx = []
    try:
        # Locate the Excel button. CRITICAL: vAuto's popup has an `in-active`
        # class on the Excel control until the data finishes loading. Clicking
        # while in-active queues but doesn't reliably fire the download (see
        # Agent B diagnosis). Poll up to 30s waiting for the active state.
        loc = None
        for attempt in range(30):
            loc = page.evaluate(JS_FIND_EXCEL_TARGET)
            if loc and loc.get('result') == 'found' and not loc.get('is_inactive'):
                break
            time.sleep(1)
            if attempt == 0 or (attempt + 1) % 5 == 0:
                print(f'  [rbook] excel target: {loc} '
                      f'(waiting for in-active to clear, attempt {attempt+1}/30)',
                      flush=True)
        print(f'  [rbook] excel target ready: {loc}', flush=True)
        if not loc or loc.get('result') != 'found':
            raise RuntimeError(f'excel target not found: {loc}')
        if loc.get('is_inactive'):
            # 30s elapsed and still in-active — proceed anyway, but warn
            print('  [rbook] WARNING — excel still in-active after 30s, '
                  'clicking anyway', flush=True)
        cx = float(loc.get('cx') or 0)
        cy = float(loc.get('cy') or 0)
        if cx <= 0 or cy <= 0:
            raise RuntimeError(f'excel target offscreen: cx={cx} cy={cy}')

        # Per user observation: vAuto sometimes opens content in a separate
        # tab/window. Listen for 'popup' events on this page AND new 'page'
        # events on the context so we can track downloads anywhere.
        ctx = page.context
        new_pages = []
        def _on_new_page(p):
            new_pages.append(p)
            print(f'  [rbook] NEW PAGE opened: {p.url[:120]}', flush=True)
        ctx.on('page', _on_new_page)

        # Wire download listeners on ALL existing + future pages so we catch
        # the file regardless of which page initiates it. CRITICAL: the
        # listener must call save_as() INLINE — vAuto's Excel generation
        # can take 30-90s, and if we wait then poll, the Download object's
        # internal stream may close. Save-on-event is the reliable path.
        import os as _os
        captured_downloads = []
        saved_paths = []  # set by _on_download when save succeeds

        _save_dir = _os.path.join(_os.path.expanduser('~'), 'rbook_downloads')
        try: _os.makedirs(_save_dir, exist_ok=True)
        except Exception: pass

        def _on_download(d):
            captured_downloads.append(d)
            fname = d.suggested_filename or 'rbook.xlsx'
            target = _os.path.join(
                _save_dir,
                f'rbook_{job.get("bid_id","x")}_{int(time.time()*1000)}_{fname}')
            print(f'  [rbook] DOWNLOAD event: {fname}', flush=True)
            try:
                d.save_as(target)
                saved_paths.append(target)
                print(f'  [rbook] saved → {target}', flush=True)
            except Exception as e:
                print(f'  [rbook] save_as failed inline: {e}', flush=True)
        for p in ctx.pages:
            try: p.on('download', _on_download)
            except Exception: pass
        ctx.on('page', lambda p: p.on('download', _on_download))

        # vAuto's Excel download bypasses Playwright's `expect_download`
        # event. Watch BOTH the CDP-configured download dir (set by
        # Browser.page() via Browser.setDownloadBehavior) AND the OS
        # Downloads dir as backup.
        import os as _os
        DL_DIRS = []
        cdp_dir = _os.path.join(_os.path.expanduser('~'), 'rbook_downloads')
        if _os.path.isdir(cdp_dir):
            DL_DIRS.append(cdp_dir)
        os_dl = _os.path.expanduser('~/Downloads')
        if _os.path.isdir(os_dl):
            DL_DIRS.append(os_dl)
        if not DL_DIRS:
            DL_DIRS = [r'C:\Users\worker-1\Downloads']

        # Match xlsx/xls/csv/temp AND bare GUID-named files (Chrome's
        # `allowAndName` mode strips extensions). UUID regex catches files
        # like 60c3eb3d-b16d-476e-96af-654907e1a24e
        import re as _re
        _UUID_RE = _re.compile(r'^[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}$', _re.I)

        def _xlsx_state():
            """Return dict of (full_path → mtime) for FINISHED downloads in
            watched dirs. We deliberately exclude in-progress markers
            (.tmp, .crdownload, bare-GUID without extension) — those rename
            mid-flight and using their path causes WinError 2."""
            out = {}
            for d in DL_DIRS:
                try:
                    for f in _os.listdir(d):
                        fl = f.lower()
                        if fl.endswith('.xlsx') or fl.endswith('.xls') or fl.endswith('.csv'):
                            full = _os.path.join(d, f)
                            try: out[full] = _os.path.getmtime(full)
                            except Exception: pass
                except Exception: pass
            return out

        before = _xlsx_state()
        print(f'  [rbook] watching dirs: {DL_DIRS} ({len(before)} existing xlsx)',
              flush=True)

        # Bring the page to front so the click lands in a focused window —
        # Chrome can defer clicks/downloads in unfocused tabs (Agent B).
        try: page.bring_to_front()
        except Exception: pass
        time.sleep(0.3)

        # Issue REAL mouse click at the element coords (user gesture)
        page.mouse.click(cx, cy)
        print(f'  [rbook] mouse.click({int(cx)}, {int(cy)}) on '
              f'{loc.get("tag")} .{(loc.get("cls") or "")[:50]}',
              flush=True)
        _shot('02_after_excel_click')

        # Single 300s wait on saved_paths. vAuto's server-side Excel
        # generation can take 60-200s before the DOWNLOAD event fires; once
        # it does, save_as() runs INLINE in the listener and appends to
        # saved_paths. So we just wait on that one signal. Filesystem
        # fallback covers the (rare) case where Playwright misses the event.
        new_path = None
        last_status = 0
        t_end = time.time() + 300
        while time.time() < t_end:
            time.sleep(1)
            if saved_paths:
                new_path = saved_paths[-1]
                print(f'  [rbook] download saved via event: {new_path}',
                      flush=True)
                break
            current = _xlsx_state()
            new_files = [f for f in current
                         if f not in before or current[f] != before.get(f)]
            if new_files:
                new_files.sort(key=lambda f: current[f], reverse=True)
                new_path = new_files[0]
                print(f'  [rbook] new xlsx detected via filesystem: {new_path}',
                      flush=True)
                break
            # Heartbeat every 30s so user can see we're still waiting
            elapsed = int(time.time() - (t_end - 300))
            if elapsed - last_status >= 30:
                print(f'  [rbook] waiting for download... {elapsed}s '
                      f'(captured={len(captured_downloads)} saved={len(saved_paths)})',
                      flush=True)
                last_status = elapsed

        if new_pages:
            print(f'  [rbook] {len(new_pages)} new page(s) opened during click. URLs:',
                  flush=True)
            for p_ in new_pages:
                try: print(f'    {p_.url[:140]}', flush=True)
                except Exception: pass

        if not new_path:
            raise RuntimeError(
                f'no xlsx after 300s · captured={len(captured_downloads)} '
                f'saved={len(saved_paths)} new_pages={len(new_pages)}')

        # Wait for size to stabilize (download complete)
        prev_sz = -1
        for _ in range(30):
            try: sz = _os.path.getsize(new_path)
            except FileNotFoundError: sz = 0
            if sz > 1000 and sz == prev_sz:
                break
            prev_sz = sz
            time.sleep(0.5)

        print(f'  [rbook] xlsx downloaded → {new_path} '
              f'({_os.path.getsize(new_path):,} bytes)', flush=True)
        rows_from_xlsx = _parse_xlsx(new_path)

        # Cleanup — delete the file so it doesn't accumulate
        try: _os.remove(new_path)
        except Exception: pass
    except Exception as _xlsx_err:
        print(f'  [rbook] excel path failed: {_xlsx_err}', flush=True)

    # Step 3: extract stocking-report + paginate through ALL popup pages.
    # vAuto's popup virtualizes ~12 rows per page, with pagination buttons
    # at the bottom. We click Next until we've drained all pages.
    raw = page.evaluate(JS_EXTRACT_RBOOK)
    stocking = (raw or {}).get('stocking_report')
    in_dom_rows = (raw or {}).get('rows') or []
    in_dom_count_text = (raw or {}).get('count_text')

    # If Excel didn't work, paginate the in-DOM popup
    if not rows_from_xlsx and in_dom_rows:
        all_rows_by_vin = {r['vin']: r for r in in_dom_rows if r.get('vin')}
        print(f'  [rbook] page 1: {len(all_rows_by_vin)} rows (paginating popup...)',
              flush=True)

        # Pagination JS: locate Next-page button, return rect for trusted click
        JS_FIND_NEXT_PAGE = r"""() => {
            function allEls(root) {
                const out = [];
                function walk(n) {
                    if (!n || !n.querySelectorAll) return;
                    try {
                        n.querySelectorAll('*').forEach(el => {
                            out.push(el);
                            if (el.shadowRoot) walk(el.shadowRoot);
                        });
                    } catch (e) {}
                }
                walk(root);
                return out;
            }
            const all = allEls(document);
            // Look for a 'Next' button - typically aria-label="Next" or class
            // contains 'next'/'page-next', or text label arrow ">"
            for (const el of all) {
                if (!el.offsetParent) continue;
                const cls = (el.className || '').toString().toLowerCase();
                const aria = (el.getAttribute?.('aria-label') || '').toLowerCase();
                const txt = (el.textContent || '').trim();
                if ((aria === 'next' || aria === 'next page' ||
                     /\bk-i-arrow-e\b|\bk-pager-next\b|\bpage-next\b/.test(cls) ||
                     (txt === '›' || txt === '>' || txt === 'Next'))
                    && (el.tagName === 'BUTTON' || el.tagName === 'A' ||
                        el.getAttribute?.('role') === 'button' ||
                        el.tagName === 'SPAN')) {
                    if (el.disabled || el.classList?.contains('k-state-disabled')
                        || el.classList?.contains('disabled')
                        || el.getAttribute?.('aria-disabled') === 'true') {
                        return {result: 'last_page'};
                    }
                    const r = el.getBoundingClientRect();
                    if (r.width <= 0 || r.height <= 0) continue;
                    el.scrollIntoView?.({block: 'center', behavior: 'instant'});
                    const r2 = el.getBoundingClientRect();
                    return {
                        result: 'found',
                        cx: r2.left + r2.width / 2,
                        cy: r2.top + r2.height / 2,
                        cls: cls.substring(0, 60),
                    };
                }
            }
            return {result: 'no_next'};
        }"""

        for pg_n in range(2, 30):  # cap at 30 pages (~600 rows max)
            next_loc = page.evaluate(JS_FIND_NEXT_PAGE)
            if not next_loc or next_loc.get('result') != 'found':
                print(f'  [rbook] pagination ended: {next_loc}', flush=True)
                break
            page.mouse.click(float(next_loc['cx']), float(next_loc['cy']))
            time.sleep(2)  # let new page render
            page_raw = page.evaluate(JS_EXTRACT_RBOOK)
            page_rows = (page_raw or {}).get('rows') or []
            new_count = 0
            for r in page_rows:
                if r.get('vin') and r['vin'] not in all_rows_by_vin:
                    all_rows_by_vin[r['vin']] = r
                    new_count += 1
            print(f'  [rbook] page {pg_n}: +{new_count} new rows '
                  f'(total {len(all_rows_by_vin)})', flush=True)
            if new_count == 0:
                # No new rows on this page — likely cycled back to start
                break
        in_dom_rows = list(all_rows_by_vin.values())

    # Pick the larger row source — xlsx wins if present and has data
    if rows_from_xlsx:
        final_rows = rows_from_xlsx
        source = 'xlsx'
    else:
        final_rows = in_dom_rows
        source = f'in_dom_paginated' if len(in_dom_rows) > 30 else 'in_dom_fallback'

    if not final_rows:
        print(f'  [rbook] no rows from either source. count_text={in_dom_count_text} '
              f'expected={expected_count}', flush=True)
        return None

    print(f'  [rbook] extracted {len(final_rows)} rows via {source} '
          f'(vAuto reports {in_dom_count_text or expected_count} total)',
          flush=True)

    return {
        'rows': final_rows,
        'count_text': in_dom_count_text or expected_count,
        'stocking_report': stocking,
        'panel_found': True,
        'n_visible': len(final_rows),
        'source': source,
    }


def _scrape_legacy_polling(page, job, expected_count):
    """Old in-DOM polling path. Kept as a function reference for diagnosis
    but not called in normal flow — Excel-export is the primary path now."""
    # poll for rows to render. Watch for VIN+$ patterns appearing
    # in the appraisal shadow root. The competitive set takes 5-20s to
    # populate after click depending on the YMM and Cox latency.
    presence_check = r"""
        () => {
            const app = document.querySelector('profit-time-guided-appraisal');
            if (!app || !app.shadowRoot) return {n: 0, len: 0};
            const html = app.shadowRoot.innerHTML || '';
            const vinM = html.match(/[A-HJ-NPR-Z0-9]{17}/g) || [];
            const uniqVins = new Set(vinM).size;
            const dollarM = html.match(/\$\d{1,3}(?:,\d{3})+/g) || [];
            return {uniq_vins: uniqVins, dollars: dollarM.length, len: html.length};
        }
    """
    saw_content = False
    for attempt in range(13):  # ~26s total
        time.sleep(2)
        try:
            chk = page.evaluate(presence_check)
            uv = chk.get('uniq_vins', 0)
            d = chk.get('dollars', 0)
            if uv >= 3 and d >= 5:
                print(f'  [rbook] content rendered (VINs={uv}, $={d}, '
                      f'len={chk.get("len", 0):,}) after {(attempt+1)*2}s',
                      flush=True)
                saw_content = True
                break
        except Exception as e:
            pass
    if not saw_content:
        print(f'  [rbook] content never rendered after 26s — '
              f'click may not have opened panel. expected count={expected_count}',
              flush=True)
        return None

    # Step 3: extract
    raw = page.evaluate(JS_EXTRACT_RBOOK)
    if not raw or raw.get('error'):
        print(f'  [rbook] extract failed: {raw}', flush=True)
        return None

    rows = raw.get('rows') or []
    stocking = raw.get('stocking_report')
    count_text = raw.get('count_text')

    if not rows:
        print(f'  [rbook] no rows after extraction. count_text={count_text} '
              f'expected={expected_count}', flush=True)
        return None

    print(f'  [rbook] extracted {len(rows)} rows '
          f'(vAuto reports {count_text or expected_count} total)', flush=True)

    return {
        'rows': rows,
        'count_text': count_text or expected_count,
        'stocking_report': stocking,
        'panel_found': True,
        'n_visible': len(rows),
    }
